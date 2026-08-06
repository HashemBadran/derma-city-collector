"""Excel export — the aged position plus the full collector's diary."""

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from db import STATUSES

STATUS_LABEL = dict(STATUSES)

THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill('solid', fgColor='1F3864')
HDR_FONT = Font(bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(bold=True, size=14, color='1F3864')
SUB_FONT = Font(italic=True, size=10, color='595959')
TOT_FILL = PatternFill('solid', fgColor='DDEBF7')
HOT_FILL = PatternFill('solid', fgColor='FCE4E4')
WARN_FILL = PatternFill('solid', fgColor='FFF2CC')
MONEY = '#,##0.00;[Red]-#,##0.00'
DATE_FMT = 'yyyy-mm-dd'


def _header(ws, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill, c.font, c.border = HDR_FILL, HDR_FONT, BORDER
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 30


def _finish(ws, freeze, filter_ref=None):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = freeze
    if filter_ref:
        ws.auto_filter.ref = filter_ref
    ws.page_setup.orientation = 'landscape'
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1


def build(customers, as_of, currency):
    wb = Workbook()

    ws = wb.active
    ws.title = 'Customers'
    ws['A1'] = 'Derma City Collector — Receivables'
    ws['A1'].font = TITLE_FONT
    ws['A2'] = f'As of {as_of}   |   Currency: {currency}   |   {len(customers)} customers'
    ws['A2'].font = SUB_FONT
    ws.merge_cells('A1:L1')
    ws.merge_cells('A2:L2')

    headers = ['#', 'Customer', 'Phone', 'City', 'Area', 'Docs', 'Oldest (days)',
               'Total Open', 'Overdue', 'Status', 'Needs Visit', 'Next Action',
               'Promise Date', 'Location Set']
    widths = [5, 44, 16, 14, 13, 8, 14, 18, 18, 16, 12, 14, 14, 12]
    _header(ws, 4, headers, widths)

    r = 5
    for idx, c in enumerate(customers, start=1):
        col = 1
        ws.cell(row=r, column=col, value=idx).alignment = Alignment(horizontal='center'); col += 1
        ws.cell(row=r, column=col, value=c['name']).alignment = Alignment(horizontal='right'); col += 1
        ws.cell(row=r, column=col, value=c['phone']); col += 1
        ws.cell(row=r, column=col, value=c['city']); col += 1
        ws.cell(row=r, column=col, value=c['area']); col += 1
        ws.cell(row=r, column=col, value=c['documents']).alignment = Alignment(horizontal='center'); col += 1

        oc = ws.cell(row=r, column=col, value=c['oldest_days'] if c['oldest_days'] > 0 else 'Not due')
        oc.alignment = Alignment(horizontal='center')
        if c['oldest_days'] >= 365:
            oc.fill = HOT_FILL
        elif c['oldest_days'] >= 180:
            oc.fill = WARN_FILL
        col += 1

        t = ws.cell(row=r, column=col, value=c['total_open']); t.number_format = MONEY; col += 1
        o = ws.cell(row=r, column=col, value=c['overdue_total']); o.number_format = MONEY
        if c['overdue_total'] > 0:
            o.font = Font(color='B3261E')
        col += 1
        ws.cell(row=r, column=col, value=STATUS_LABEL.get(c['status'], c['status'])).alignment = \
            Alignment(horizontal='center'); col += 1
        nv = ws.cell(row=r, column=col, value='Yes' if c['needs_visit'] else '')
        nv.alignment = Alignment(horizontal='center')
        if c['needs_visit']:
            nv.fill = WARN_FILL
        col += 1
        na = ws.cell(row=r, column=col, value=c['next_action_date'] or '')
        na.number_format = DATE_FMT; na.alignment = Alignment(horizontal='center'); col += 1
        pd_ = ws.cell(row=r, column=col, value=c['promise_date'] or '')
        pd_.number_format = DATE_FMT; pd_.alignment = Alignment(horizontal='center'); col += 1
        loc = ws.cell(row=r, column=col,
                      value='Yes' if c.get('lat') is not None else '')
        loc.alignment = Alignment(horizontal='center')

        for j in range(1, len(headers) + 1):
            ws.cell(row=r, column=j).border = BORDER
        r += 1

    lbl = ws.cell(row=r, column=2, value=f'TOTAL — {len(customers)} customers')
    lbl.font, lbl.fill = Font(bold=True, size=11), TOT_FILL
    total_open = round(sum(c['total_open'] for c in customers), 2)
    total_overdue = round(sum(c['overdue_total'] for c in customers), 2)
    t = ws.cell(row=r, column=8, value=total_open)
    t.number_format, t.font, t.fill = MONEY, Font(bold=True, size=12), TOT_FILL
    o = ws.cell(row=r, column=9, value=total_overdue)
    o.number_format, o.font, o.fill = MONEY, Font(bold=True), TOT_FILL
    for j in range(1, len(headers) + 1):
        ws.cell(row=r, column=j).border = BORDER
        if j not in (8, 9, 2):
            ws.cell(row=r, column=j).fill = TOT_FILL
    _finish(ws, 'A5', f'A4:{get_column_letter(len(headers))}{r - 1}')
    return wb


def visits_sheet(wb, conn):
    rows = conn.execute(
        'SELECT v.created_at, c.name AS customer, v.channel, v.status, v.outcome,'
        '       v.promise_date, v.promise_amount, v.next_action_date, v.notes,'
        '       ct.name AS contact_name, ct.position AS contact_position'
        '  FROM visits v'
        '  LEFT JOIN customers c ON c.partner_id = v.partner_id'
        '  LEFT JOIN contacts ct ON ct.id = v.contact_id'
        ' ORDER BY v.created_at DESC', []
    ).fetchall()
    if not rows:
        return wb
    ws = wb.create_sheet('Visit Log')
    ws['A1'] = 'Visit / Call Log'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:K1')
    _header(ws, 3, ['When', 'Customer', 'Channel', 'Contact', 'Position', 'Status',
                    'Outcome', 'Promise Date', 'Promise Amount', 'Next Action', 'Notes'],
            [18, 36, 12, 22, 20, 14, 30, 14, 16, 14, 50])
    r = 4
    for row in rows:
        vals = [row['created_at'], row['customer'] or '', row['channel'],
                row['contact_name'] or '', row['contact_position'] or '',
                STATUS_LABEL.get(row['status'], row['status']), row['outcome'],
                row['promise_date'] or '', row['promise_amount'] or 0,
                row['next_action_date'] or '', row['notes']]
        for j, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=j, value=v)
            cell.border = BORDER
            cell.alignment = Alignment(vertical='top', wrap_text=(j in (7, 11)))
            if j in (8, 10):
                cell.number_format = DATE_FMT
            if j == 9:
                cell.number_format = MONEY
        r += 1
    _finish(ws, 'A4', f'A3:K{r - 1}')
    return wb


def reconciliations_sheet(wb, conn):
    rows = conn.execute(
        'SELECT r.reconciled_date, c.name AS customer, r.amount, r.signed_by,'
        '       r.signed_position, r.notes'
        '  FROM reconciliations r LEFT JOIN customers c ON c.partner_id = r.partner_id'
        ' ORDER BY r.reconciled_date DESC', []
    ).fetchall()
    if not rows:
        return wb
    ws = wb.create_sheet('Signed Reconciliations')
    ws['A1'] = 'Signed Account Reconciliations'
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A1:F1')
    _header(ws, 3, ['Date', 'Customer', 'Amount', 'Signed By', 'Position', 'Notes'],
            [14, 40, 18, 26, 20, 50])
    r = 4
    for row in rows:
        vals = [row['reconciled_date'], row['customer'] or '', row['amount'],
                row['signed_by'], row['signed_position'], row['notes']]
        for j, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=j, value=v)
            cell.border = BORDER
            if j == 1:
                cell.number_format = DATE_FMT
            if j == 3:
                cell.number_format = MONEY
        r += 1
    _finish(ws, 'A4', f'A3:F{r - 1}')
    return wb


def to_bytes(customers, as_of, currency, conn):
    wb = build(customers, as_of, currency)
    visits_sheet(wb, conn)
    reconciliations_sheet(wb, conn)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
